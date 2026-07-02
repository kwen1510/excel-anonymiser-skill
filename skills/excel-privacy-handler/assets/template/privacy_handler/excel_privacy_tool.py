#!/usr/bin/env python3
"""Locked local-first Excel privacy handler."""

from __future__ import annotations

import argparse
import base64
import csv
import datetime as dt
import getpass
import hashlib
import io
import json
import re
import secrets
import sys
import uuid
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

try:
    from cryptography.fernet import Fernet
except Exception:  # pragma: no cover - dependency availability is environment-specific.
    Fernet = None

try:
    import msoffcrypto
except Exception:  # pragma: no cover - dependency availability is environment-specific.
    msoffcrypto = None


TOOL_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = TOOL_DIR.parent
LOCKED_FILES = [
    TOOL_DIR / "excel_privacy_tool.py",
    TOOL_DIR / "privacy_config.schema.json",
    TOOL_DIR / "README.md",
]
APPROVED_CHECKSUMS = TOOL_DIR / "approved_checksums.json"
MODES = {"keep", "uuid", "stable_hash", "encrypt", "blank"}
KEEP_OPTIONS = {"first", "last"}
HEADER_NAME_RE = re.compile(r"^[^\r\n]+$")


class PrivacyToolError(Exception):
    """Safe CLI error. Message must not contain original workbook values."""


def safe_exit(message: str, code: int = 1) -> None:
    print(f"Error: {message}", file=sys.stderr)
    raise SystemExit(code)


def now_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).astimezone().isoformat(timespec="seconds")


def resolve_path(path: str | Path) -> Path:
    candidate = Path(path)
    if not candidate.is_absolute():
        candidate = PROJECT_ROOT / candidate
    return candidate.resolve()


def relative_display(path: str | Path) -> str:
    resolved = resolve_path(path)
    try:
        return str(resolved.relative_to(PROJECT_ROOT))
    except ValueError:
        return str(resolved)


def load_json(path: str | Path) -> dict[str, Any]:
    try:
        with open(resolve_path(path), "r", encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as exc:
        raise PrivacyToolError(f"Invalid JSON in {relative_display(path)} at line {exc.lineno}") from exc
    if not isinstance(data, dict):
        raise PrivacyToolError(f"Config must be a JSON object: {relative_display(path)}")
    return data


def is_within(child: Path, parent: Path) -> bool:
    try:
        child.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def validate_output_safety(config: dict[str, Any], config_kind: str = "privacy") -> None:
    output = config.get("output", {})
    safe_dir = resolve_path(output.get("safe_dir", "safe_for_codex"))
    private_key_dir = resolve_path(output.get("private_key_dir", "private_keys"))
    restored_dir = resolve_path(output.get("restored_dir", "restored_outputs"))

    if safe_dir == private_key_dir or is_within(private_key_dir, safe_dir):
        raise PrivacyToolError("Private key directory must not be inside safe_for_codex")
    if safe_dir == restored_dir or is_within(restored_dir, safe_dir):
        raise PrivacyToolError("Restored output directory must not be inside safe_for_codex")

    key_file = config.get("key_file")
    if isinstance(key_file, dict):
        secret = key_file.get("secret_key_file")
        if secret and is_within(resolve_path(secret), safe_dir):
            raise PrivacyToolError("Secret key file must not be inside safe_for_codex")

    if config_kind == "restore":
        output_file = config.get("output_file")
        if output_file and is_within(resolve_path(output_file), safe_dir):
            raise PrivacyToolError("Restored output file must not be inside safe_for_codex")


def validate_config_data(config: dict[str, Any], kind_hint: str | None = None) -> str:
    if "sheet" in config and "by" in config:
        kind = "dedupe"
    elif "key_file" in config and "output_file" in config and "sheets" not in config:
        kind = "restore"
    else:
        kind = kind_hint or "privacy"

    if kind == "dedupe":
        sheet = config.get("sheet")
        by = config.get("by")
        keep = config.get("keep", "first")
        output_file = config.get("output_file")
        if not isinstance(sheet, str) or not sheet.strip():
            raise PrivacyToolError("Dedupe config requires a sheet name")
        if not isinstance(by, list) or not by or not all(isinstance(x, str) and x.strip() for x in by):
            raise PrivacyToolError("Dedupe config requires a non-empty by list")
        if keep not in KEEP_OPTIONS:
            raise PrivacyToolError("Dedupe keep must be first or last")
        if not isinstance(output_file, str) or not output_file.strip():
            raise PrivacyToolError("Dedupe config requires output_file")
        if is_within(resolve_path(output_file), resolve_path("private_keys")):
            raise PrivacyToolError("Safe dedupe output must not be written to private_keys")
        return kind

    if kind == "restore":
        key_file = config.get("key_file")
        output_file = config.get("output_file")
        if not isinstance(key_file, str) or not key_file.strip():
            raise PrivacyToolError("Restore config requires key_file")
        if not isinstance(output_file, str) or not output_file.strip():
            raise PrivacyToolError("Restore config requires output_file")
        validate_output_safety({"output": {}, "output_file": output_file}, "restore")
        if is_within(resolve_path(key_file), resolve_path("safe_for_codex")):
            raise PrivacyToolError("Privacy key file must not be inside safe_for_codex")
        return kind

    required = ["header_row", "output", "key_file", "sheets"]
    for field in required:
        if field not in config:
            raise PrivacyToolError(f"Privacy config missing required field: {field}")
    if not isinstance(config["header_row"], int) or config["header_row"] < 1:
        raise PrivacyToolError("header_row must be a positive integer")
    if not isinstance(config["output"], dict):
        raise PrivacyToolError("output must be an object")
    for field in ["safe_dir", "private_key_dir", "restored_dir"]:
        if not isinstance(config["output"].get(field), str) or not config["output"].get(field).strip():
            raise PrivacyToolError(f"output.{field} is required")
    if not isinstance(config["key_file"], dict):
        raise PrivacyToolError("key_file must be an object")
    if config["key_file"].get("format") != "xlsx":
        raise PrivacyToolError("key_file.format must be xlsx")
    if not isinstance(config["key_file"].get("encrypt_original_values"), bool):
        raise PrivacyToolError("key_file.encrypt_original_values must be boolean")
    if not isinstance(config["sheets"], dict) or not config["sheets"]:
        raise PrivacyToolError("sheets must be a non-empty object")

    for sheet_name, sheet_config in config["sheets"].items():
        if not isinstance(sheet_name, str) or not HEADER_NAME_RE.match(sheet_name):
            raise PrivacyToolError("Configured sheet names must be non-empty single-line strings")
        columns = sheet_config.get("columns") if isinstance(sheet_config, dict) else None
        if not isinstance(columns, dict):
            raise PrivacyToolError(f"Sheet {sheet_name} requires columns")
        for header, column_config in columns.items():
            if not isinstance(header, str) or not HEADER_NAME_RE.match(header):
                raise PrivacyToolError(f"Configured column names in sheet {sheet_name} must be single-line strings")
            mode = column_config.get("mode") if isinstance(column_config, dict) else None
            if mode not in MODES:
                raise PrivacyToolError(f"Invalid mode for sheet {sheet_name}, column {header}")
            prefix = column_config.get("prefix", header.upper().replace(" ", "_"))
            if not isinstance(prefix, str) or not prefix.strip():
                raise PrivacyToolError(f"Invalid prefix for sheet {sheet_name}, column {header}")

    validate_output_safety(config, "privacy")
    return kind


def prompt_password_popup(workbook_name: str) -> str | None:
    try:
        import tkinter as tk
        from tkinter import simpledialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        password = simpledialog.askstring(
            "Excel workbook password",
            f"Enter password for {workbook_name}:",
            show="*",
            parent=root,
        )
        root.destroy()
        return password
    except Exception:
        return getpass.getpass(f"Enter password for {workbook_name}: ")


def _openpyxl_load(source: Path | io.BytesIO):
    return load_workbook(source)


def decrypt_encrypted_workbook_bytes(path: Path, password_provider: Callable[[str], str | None]) -> io.BytesIO:
    # PRIVACY-SENSITIVE:
    # This function handles the workbook password locally.
    # Do not print, log, return, or expose the password.
    if msoffcrypto is None:
        raise PrivacyToolError("Workbook appears encrypted; install msoffcrypto-tool to use password popup support")
    password = password_provider(path.name)
    if not password:
        raise PrivacyToolError("Workbook password was not provided")
    decrypted = io.BytesIO()
    try:
        with open(path, "rb") as f:
            office_file = msoffcrypto.OfficeFile(f)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
    except Exception as exc:
        raise PrivacyToolError("Unable to decrypt workbook; password may be incorrect") from exc
    decrypted.seek(0)
    return decrypted


def load_workbook_private(path: Path, password_provider: Callable[[str], str | None] = prompt_password_popup):
    # PRIVACY-SENSITIVE:
    # This function reads original workbook contents locally.
    # Do not print, log, return, or expose raw values.
    try:
        return _openpyxl_load(path)
    except Exception as first_exc:
        decrypted = decrypt_encrypted_workbook_bytes(path, password_provider)
        try:
            return _openpyxl_load(decrypted)
        except Exception as exc:
            raise PrivacyToolError(f"Unable to open workbook {path.name}") from exc or first_exc


def safe_header_value(value: Any, column_index: int) -> str:
    if value is None:
        return f"Column_{column_index}"
    text = str(value)
    if "\r" in text or "\n" in text:
        return text.replace("\r", " ").replace("\n", " ")
    return text


def get_headers(ws, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[header_row]:
        header = safe_header_value(cell.value, cell.column)
        headers[header] = cell.column
    return headers


def cmd_inspect(args: argparse.Namespace) -> None:
    workbook_path = resolve_path(args.input)
    wb = load_workbook_private(workbook_path)
    sheets = []
    for ws in wb.worksheets:
        headers = []
        for cell in ws[args.header_row]:
            headers.append(
                {
                    "column_index": cell.column,
                    "column_letter": get_column_letter(cell.column),
                    "header": safe_header_value(cell.value, cell.column),
                }
            )
        sheets.append(
            {
                "sheet_name": ws.title,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "headers": headers,
            }
        )
    print(json.dumps({"workbook": workbook_path.name, "header_row": args.header_row, "sheets": sheets}, indent=2))


def replacement_uuid(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:8]}"


def replacement_hash(prefix: str, salt: str, value: Any) -> str:
    digest = hashlib.sha256(f"{salt}\0{str(value)}".encode("utf-8")).hexdigest()[:12]
    return f"{prefix}_{digest}"


def require_fernet() -> Any:
    if Fernet is None:
        raise PrivacyToolError("cryptography is required for encrypted key files or encrypt mode")
    return Fernet


def encrypt_value(fernet: Any, value: Any) -> str:
    return fernet.encrypt(str(value).encode("utf-8")).decode("ascii")


def decrypt_value(fernet: Any, value: str) -> str:
    return fernet.decrypt(value.encode("ascii")).decode("utf-8")


def make_secret_key(path: Path | None) -> tuple[Any | None, Path | None]:
    if path is None:
        return None, None
    fernet_cls = require_fernet()
    path.parent.mkdir(parents=True, exist_ok=True)
    key = fernet_cls.generate_key()
    path.write_bytes(key)
    return fernet_cls(key), path


def load_secret_key(path: Path | None) -> Any | None:
    if path is None:
        return None
    fernet_cls = require_fernet()
    return fernet_cls(path.read_bytes())


def configured_output_paths(input_path: Path, config: dict[str, Any]) -> tuple[Path, Path, Path, Path | None]:
    output = config["output"]
    safe_dir = resolve_path(output["safe_dir"])
    private_dir = resolve_path(output["private_key_dir"])
    stem = input_path.stem
    anonymised = safe_dir / f"{stem}.anonymised.xlsx"
    manifest = safe_dir / f"{stem}.privacy_manifest.json"
    key_file = private_dir / f"{stem}.privacy_key.xlsx"
    secret_cfg = config.get("key_file", {}).get("secret_key_file")
    secret = resolve_path(secret_cfg) if secret_cfg else private_dir / f"{stem}.secret.key"
    if not config.get("key_file", {}).get("encrypt_original_values") and not uses_encrypt_mode(config):
        secret = None
    return anonymised, manifest, key_file, secret


def uses_encrypt_mode(config: dict[str, Any]) -> bool:
    for sheet_config in config.get("sheets", {}).values():
        for column_config in sheet_config.get("columns", {}).values():
            if column_config.get("mode") == "encrypt":
                return True
    return False


def check_not_overwrite_source(output_path: Path, source_path: Path) -> None:
    if output_path.resolve() == source_path.resolve():
        raise PrivacyToolError("Refusing to overwrite the original workbook")


def anonymise_headers(ws, configured_columns: dict[str, Any], header_row: int) -> dict[str, str]:
    header_map = {}
    for cell in ws[header_row]:
        header = safe_header_value(cell.value, cell.column)
        if header in configured_columns:
            new_header = f"HEADER_{uuid.uuid4().hex[:8]}"
            header_map[header] = new_header
            cell.value = new_header
    return header_map


def cmd_anonymise(args: argparse.Namespace) -> None:
    verify_self_before_sensitive()
    input_path = resolve_path(args.input)
    config = load_json(args.config)
    validate_config_data(config, "privacy")
    anonymised_path, manifest_path, key_path, secret_path = configured_output_paths(input_path, config)
    for output_path in [anonymised_path, manifest_path, key_path] + ([secret_path] if secret_path else []):
        check_not_overwrite_source(output_path, input_path)
    validate_output_safety(config, "privacy")

    wb = load_workbook_private(input_path)
    fernet, actual_secret_path = make_secret_key(secret_path)
    header_row = config["header_row"]
    mappings: dict[tuple[str, str], dict[Any, str]] = {}
    mapping_rows: list[dict[str, Any]] = []
    manifest: dict[str, Any] = {
        "source_workbook_name": input_path.name,
        "created_at": now_iso(),
        "anonymised_workbook": relative_display(anonymised_path),
        "privacy_key_location": relative_display(key_path),
        "sheets": {},
    }

    for original_sheet_name, sheet_config in config["sheets"].items():
        if original_sheet_name not in wb.sheetnames:
            raise PrivacyToolError(f"Configured sheet not found: {original_sheet_name}")
        ws = wb[original_sheet_name]
        headers = get_headers(ws, header_row)
        configured_columns = sheet_config["columns"]
        missing = [header for header in configured_columns if header not in headers]
        if missing:
            raise PrivacyToolError(f"Missing configured column in sheet {original_sheet_name}: {missing[0]}")

        if config.get("anonymise_sheet_names"):
            safe_sheet_name = f"SHEET_{uuid.uuid4().hex[:8]}"
            ws.title = safe_sheet_name
        else:
            safe_sheet_name = original_sheet_name

        header_name_map: dict[str, str] = {}
        if config.get("anonymise_headers"):
            for header in configured_columns:
                header_name_map[header] = f"HEADER_{uuid.uuid4().hex[:8]}"

        sheet_manifest = {"rows": ws.max_row, "columns": ws.max_column, "anonymised_columns": []}
        salt = secrets.token_hex(16)
        for header, column_config in configured_columns.items():
            mode = column_config["mode"]
            prefix = column_config.get("prefix", header.upper().replace(" ", "_"))
            column_index = headers[header]
            key = (original_sheet_name, header)
            mappings[key] = {}
            unique_count = 0

            if mode != "keep":
                for row_index in range(header_row + 1, ws.max_row + 1):
                    cell = ws.cell(row=row_index, column=column_index)
                    original = cell.value
                    if original is None or original == "":
                        continue
                    if mode == "blank":
                        replacement = None
                    elif mode == "uuid":
                        if original not in mappings[key]:
                            mappings[key][original] = replacement_uuid(prefix)
                        replacement = mappings[key][original]
                    elif mode == "stable_hash":
                        if original not in mappings[key]:
                            mappings[key][original] = replacement_hash(prefix, salt, original)
                        replacement = mappings[key][original]
                    elif mode == "encrypt":
                        if fernet is None:
                            raise PrivacyToolError("encrypt mode requires a secret key")
                        if original not in mappings[key]:
                            mappings[key][original] = encrypt_value(fernet, original)
                        replacement = mappings[key][original]
                    else:
                        raise PrivacyToolError(f"Invalid mode for sheet {original_sheet_name}, column {header}")
                    if original not in [row.get("_original") for row in []]:
                        pass
                    cell.value = replacement
                unique_count = len(mappings[key])

                for original, replacement in mappings[key].items():
                    key_original = encrypt_value(fernet, original) if config["key_file"].get("encrypt_original_values") else original
                    mapping_rows.append(
                        {
                            "sheet_name": original_sheet_name,
                            "target_sheet_name": safe_sheet_name,
                            "column_header": header,
                            "target_column_header": header_name_map.get(header, header),
                            "replacement_id": replacement,
                            "original_value": key_original,
                            "mode": mode,
                            "created_at": now_iso(),
                        }
                    )
                sheet_manifest["anonymised_columns"].append(
                    {"header": header, "mode": mode, "unique_count": unique_count}
                )
        if config.get("anonymise_headers"):
            for cell in ws[header_row]:
                header = safe_header_value(cell.value, cell.column)
                if header in header_name_map:
                    cell.value = header_name_map[header]
        manifest["sheets"][safe_sheet_name] = sheet_manifest

    anonymised_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    key_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(anonymised_path)
    write_manifest(manifest_path, manifest)
    write_key_workbook(key_path, mapping_rows, actual_secret_path)

    print(f"Created anonymised workbook: {relative_display(anonymised_path)}")
    print(f"Created privacy key: {relative_display(key_path)}")
    if actual_secret_path:
        print(f"Created secret key: {relative_display(actual_secret_path)}")
    print(f"Created manifest: {relative_display(manifest_path)}")
    print("Mappings created:")
    for sheet_name, sheet_config in config["sheets"].items():
        for header, column_config in sheet_config["columns"].items():
            if column_config["mode"] != "keep":
                print(f"- {sheet_name} / {header}: {len(mappings[(sheet_name, header)])} unique values")


def parse_columns(columns: str | None) -> list[str]:
    if not columns:
        raise PrivacyToolError("Pass columns to mask with --columns, for example: --columns Name,Email,Class")
    parsed = [column.strip() for column in columns.split(",") if column.strip()]
    if not parsed:
        raise PrivacyToolError("No mask columns were provided")
    return parsed


def simple_output_paths(input_path: Path, output: str | None, key_file: str | None, secret_key_file: str | None) -> tuple[Path, Path, Path]:
    masked_path = resolve_path(output) if output else resolve_path("safe_for_codex") / f"{input_path.stem}.masked.xlsx"
    key_path = resolve_path(key_file) if key_file else resolve_path("private_keys") / f"{input_path.stem}.mask_key.csv"
    secret_path = (
        resolve_path(secret_key_file)
        if secret_key_file
        else resolve_path("private_keys") / f"{input_path.stem}.mask_secret.key"
    )
    if is_within(key_path, resolve_path("safe_for_codex")):
        raise PrivacyToolError("Mask key CSV must not be written to safe_for_codex")
    if is_within(secret_path, resolve_path("safe_for_codex")):
        raise PrivacyToolError("Mask secret key must not be written to safe_for_codex")
    if is_within(masked_path, resolve_path("private_keys")):
        raise PrivacyToolError("Masked workbook must not be written to private_keys")
    check_not_overwrite_source(masked_path, input_path)
    return masked_path, key_path, secret_path


def write_simple_key_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "sheet_name",
                "column_header",
                "masked_value",
                "encrypted_original_value",
                "created_at",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)


def read_simple_key_csv(path: Path, fernet: Any) -> dict[tuple[str, str], dict[str, str]]:
    # PRIVACY-SENSITIVE:
    # This function reads private mask mappings locally.
    # Do not print, log, return to CLI output, or expose original values.
    mappings: dict[tuple[str, str], dict[str, str]] = {}
    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        required = {"sheet_name", "column_header", "masked_value", "encrypted_original_value"}
        if not required.issubset(set(reader.fieldnames or [])):
            raise PrivacyToolError("Invalid mask key CSV")
        for row in reader:
            sheet_name = row["sheet_name"]
            column_header = row["column_header"]
            masked_value = row["masked_value"]
            original = decrypt_value(fernet, row["encrypted_original_value"])
            mappings.setdefault((sheet_name, column_header), {})[masked_value] = original
    return mappings


def cmd_mask(args: argparse.Namespace) -> None:
    verify_self_before_sensitive()
    input_path = resolve_path(args.input)
    columns = parse_columns(args.columns)
    masked_path, key_path, secret_path = simple_output_paths(input_path, args.output, args.key_file, args.secret_key_file)
    fernet_cls = require_fernet()
    secret_path.parent.mkdir(parents=True, exist_ok=True)
    secret = fernet_cls.generate_key()
    secret_path.write_bytes(secret)
    fernet = fernet_cls(secret)

    wb = load_workbook_private(input_path)
    target_sheets = [args.sheet] if args.sheet else wb.sheetnames
    key_rows: list[dict[str, Any]] = []
    summary: list[tuple[str, str, int]] = []

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            raise PrivacyToolError(f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        headers = get_headers(ws, args.header_row)
        missing = [column for column in columns if column not in headers]
        if missing:
            raise PrivacyToolError(f"Missing mask column in sheet {sheet_name}: {missing[0]}")
        for column_header in columns:
            column_index = headers[column_header]
            replacements: dict[Any, str] = {}
            for row_index in range(args.header_row + 1, ws.max_row + 1):
                cell = ws.cell(row=row_index, column=column_index)
                original = cell.value
                if original is None or original == "":
                    continue
                if original not in replacements:
                    replacements[original] = f"MASK_{uuid.uuid4().hex[:12]}"
                cell.value = replacements[original]
            for original, masked_value in replacements.items():
                key_rows.append(
                    {
                        "sheet_name": sheet_name,
                        "column_header": column_header,
                        "masked_value": masked_value,
                        "encrypted_original_value": encrypt_value(fernet, original),
                        "created_at": now_iso(),
                    }
                )
            summary.append((sheet_name, column_header, len(replacements)))

    masked_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(masked_path)
    write_simple_key_csv(key_path, key_rows)
    print(f"Created masked workbook: {relative_display(masked_path)}")
    print(f"Created private mask key CSV: {relative_display(key_path)}")
    print(f"Created private secret key: {relative_display(secret_path)}")
    print("Masked columns:")
    for sheet_name, column_header, count in summary:
        print(f"- {sheet_name} / {column_header}: {count} unique values")


def cmd_unmask(args: argparse.Namespace) -> None:
    verify_self_before_sensitive()
    input_path = resolve_path(args.input)
    output_path = resolve_path(args.output) if args.output else resolve_path("restored_outputs") / f"{input_path.stem}.unmasked.xlsx"
    key_path = resolve_path(args.key_file)
    secret_path = resolve_path(args.secret_key_file)
    if is_within(output_path, resolve_path("safe_for_codex")):
        raise PrivacyToolError("Unmasked output must not be written to safe_for_codex")
    check_not_overwrite_source(output_path, input_path)
    fernet = load_secret_key(secret_path)
    mappings = read_simple_key_csv(key_path, fernet)
    wb = load_workbook_private(input_path)
    summary: list[tuple[str, str, int]] = []

    for (sheet_name, column_header), replacements in mappings.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = get_headers(ws, args.header_row)
        if column_header not in headers:
            continue
        column_index = headers[column_header]
        count = 0
        for row_index in range(args.header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row_index, column=column_index)
            if cell.value in replacements:
                cell.value = replacements[cell.value]
                count += 1
        summary.append((sheet_name, column_header, count))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Created unmasked workbook: {relative_display(output_path)}")
    print("Unmasked columns:")
    for sheet_name, column_header, count in summary:
        print(f"- {sheet_name} / {column_header}: {count} replacements")


def write_manifest(path: Path, manifest: dict[str, Any]) -> None:
    path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")


def safe_sheet_title(title: str, existing: set[str]) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", title)[:31] or "Sheet"
    base = cleaned
    i = 1
    while cleaned in existing:
        suffix = f"_{i}"
        cleaned = f"{base[:31 - len(suffix)]}{suffix}"
        i += 1
    existing.add(cleaned)
    return cleaned


def write_key_workbook(path: Path, rows: list[dict[str, Any]], secret_path: Path | None) -> None:
    wb = Workbook()
    manifest_ws = wb.active
    manifest_ws.title = "_manifest"
    manifest_ws.append(["created_at", "secret_key_file"])
    manifest_ws.append([now_iso(), relative_display(secret_path) if secret_path else None])
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault((row["sheet_name"], row["column_header"]), []).append(row)
    existing = {"_manifest"}
    for (sheet_name, header), group in grouped.items():
        ws = wb.create_sheet(safe_sheet_title(f"{sheet_name}__{header}", existing))
        ws.append(
            [
                "sheet_name",
                "target_sheet_name",
                "column_header",
                "target_column_header",
                "replacement_id",
                "original_value",
                "mode",
                "created_at",
            ]
        )
        for row in group:
            ws.append(
                [
                    row["sheet_name"],
                    row["target_sheet_name"],
                    row["column_header"],
                    row["target_column_header"],
                    row["replacement_id"],
                    row["original_value"],
                    row["mode"],
                    row["created_at"],
                ]
            )
    wb.save(path)


def read_key_workbook(path: Path, secret_key_file: Path | None) -> dict[tuple[str, str], dict[str, Any]]:
    # PRIVACY-SENSITIVE:
    # This function reads privacy key mappings locally.
    # Do not print, log, return to CLI output, or expose original values.
    wb = load_workbook(path, data_only=True)
    fernet = load_secret_key(secret_key_file)
    mappings: dict[tuple[str, str], dict[str, Any]] = {}
    for ws in wb.worksheets:
        if ws.title == "_manifest":
            continue
        header_row = [cell.value for cell in ws[1]]
        indexes = {name: idx + 1 for idx, name in enumerate(header_row)}
        required = ["sheet_name", "column_header", "replacement_id", "original_value", "mode"]
        if not all(name in indexes for name in required):
            raise PrivacyToolError(f"Invalid privacy key sheet: {ws.title}")
        for row_idx in range(2, ws.max_row + 1):
            sheet_name = ws.cell(row_idx, indexes["sheet_name"]).value
            column_header = ws.cell(row_idx, indexes["column_header"]).value
            target_sheet_name = (
                ws.cell(row_idx, indexes["target_sheet_name"]).value
                if "target_sheet_name" in indexes
                else sheet_name
            )
            target_column_header = (
                ws.cell(row_idx, indexes["target_column_header"]).value
                if "target_column_header" in indexes
                else column_header
            )
            replacement_id = ws.cell(row_idx, indexes["replacement_id"]).value
            original_value = ws.cell(row_idx, indexes["original_value"]).value
            mode = ws.cell(row_idx, indexes["mode"]).value
            if fernet is not None and original_value is not None:
                original_value = decrypt_value(fernet, str(original_value))
            mappings.setdefault((str(target_sheet_name), str(target_column_header)), {})[replacement_id] = {
                "original": original_value,
                "mode": mode,
            }
    return mappings


def cmd_restore(args: argparse.Namespace) -> None:
    verify_self_before_sensitive()
    input_path = resolve_path(args.input)
    config = load_json(args.config)
    validate_config_data(config, "restore")
    output_file = resolve_path(config["output_file"])
    key_file = resolve_path(config["key_file"])
    secret_key_file = resolve_path(config["secret_key_file"]) if config.get("secret_key_file") else None
    check_not_overwrite_source(output_file, input_path)

    wb = load_workbook_private(input_path)
    mappings = read_key_workbook(key_file, secret_key_file)
    restored_counts: dict[tuple[str, str], int] = {}

    for (sheet_name, column_header), replacements in mappings.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = get_headers(ws, 1)
        if column_header not in headers:
            continue
        col = headers[column_header]
        count = 0
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row_idx, col).value
            if value in replacements:
                ws.cell(row_idx, col).value = replacements[value]["original"]
                count += 1
        restored_counts[(sheet_name, column_header)] = count

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    print(f"Restored workbook written to {relative_display(output_file)}")
    print("Restored values:")
    for (sheet_name, column_header), count in restored_counts.items():
        print(f"- {sheet_name} / {column_header}: {count} replacements")


def cmd_dedupe(args: argparse.Namespace) -> None:
    verify_self_before_sensitive()
    input_path = resolve_path(args.input)
    config = load_json(args.config)
    validate_config_data(config, "dedupe")
    output_file = resolve_path(config["output_file"])
    check_not_overwrite_source(output_file, input_path)
    if is_within(output_file, resolve_path("private_keys")):
        raise PrivacyToolError("Dedupe output must not be written to private_keys")

    wb = load_workbook_private(input_path)
    sheet_name = config["sheet"]
    if sheet_name not in wb.sheetnames:
        raise PrivacyToolError(f"Dedupe sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    headers = get_headers(ws, 1)
    missing = [header for header in config["by"] if header not in headers]
    if missing:
        raise PrivacyToolError(f"Missing dedupe column in sheet {sheet_name}: {missing[0]}")
    cols = [headers[header] for header in config["by"]]

    rows_before = max(ws.max_row - 1, 0)
    seen: dict[tuple[Any, ...], int] = {}
    rows_to_delete: set[int] = set()
    data_rows = list(range(2, ws.max_row + 1))
    for row_idx in data_rows:
        key = tuple(ws.cell(row_idx, col).value for col in cols)
        if key in seen:
            if config.get("keep", "first") == "first":
                rows_to_delete.add(row_idx)
            else:
                rows_to_delete.add(seen[key])
                seen[key] = row_idx
        else:
            seen[key] = row_idx
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx, 1)
    rows_after = max(ws.max_row - 1, 0)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    print(f"Deduplicated sheet {sheet_name}")
    print(f"Rows before: {rows_before}")
    print(f"Rows after: {rows_after}")
    print(f"Removed duplicate rows: {rows_before - rows_after}")
    print(f"Output written to {relative_display(output_file)}")


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def current_checksums() -> dict[str, str]:
    return {relative_display(path): sha256_file(path) for path in LOCKED_FILES}


def cmd_checksum(_args: argparse.Namespace) -> None:
    print(json.dumps(current_checksums(), indent=2))


def cmd_write_approved_checksums(_args: argparse.Namespace) -> None:
    APPROVED_CHECKSUMS.write_text(json.dumps(current_checksums(), indent=2), encoding="utf-8")
    print(f"Wrote approved checksums: {relative_display(APPROVED_CHECKSUMS)}")


def verify_self_before_sensitive() -> None:
    if not APPROVED_CHECKSUMS.exists():
        raise PrivacyToolError("approved_checksums.json is missing; run write-approved-checksums after approval")
    approved = json.loads(APPROVED_CHECKSUMS.read_text(encoding="utf-8") or "{}")
    if not approved:
        raise PrivacyToolError("approved_checksums.json is empty; run write-approved-checksums after approval")
    current = current_checksums()
    if current != approved:
        raise PrivacyToolError("Privacy handler verification failed; locked files differ from approved checksums")


def cmd_verify_self(_args: argparse.Namespace) -> None:
    verify_self_before_sensitive()
    print("Privacy handler verification passed")


def cmd_validate_config(args: argparse.Namespace) -> None:
    config = load_json(args.config)
    kind = validate_config_data(config)
    print(f"Config valid: {relative_display(args.config)} ({kind})")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Locked local-first Excel privacy handler")
    sub = parser.add_subparsers(dest="command", required=True)

    inspect_p = sub.add_parser("inspect")
    inspect_p.add_argument("input")
    inspect_p.add_argument("--header-row", type=int, default=1)
    inspect_p.set_defaults(func=cmd_inspect)

    anonymise_p = sub.add_parser("anonymise")
    anonymise_p.add_argument("input")
    anonymise_p.add_argument("--config", required=True)
    anonymise_p.set_defaults(func=cmd_anonymise)

    mask_p = sub.add_parser("mask")
    mask_p.add_argument("input")
    mask_p.add_argument("--columns", required=True, help="Comma-separated headers to mask, e.g. Name,Email,Class")
    mask_p.add_argument("--sheet", help="Optional single sheet name. Defaults to all sheets.")
    mask_p.add_argument("--header-row", type=int, default=1)
    mask_p.add_argument("--output")
    mask_p.add_argument("--key-file")
    mask_p.add_argument("--secret-key-file")
    mask_p.set_defaults(func=cmd_mask)

    unmask_p = sub.add_parser("unmask")
    unmask_p.add_argument("input")
    unmask_p.add_argument("--key-file", required=True)
    unmask_p.add_argument("--secret-key-file", required=True)
    unmask_p.add_argument("--header-row", type=int, default=1)
    unmask_p.add_argument("--output")
    unmask_p.set_defaults(func=cmd_unmask)

    dedupe_p = sub.add_parser("dedupe")
    dedupe_p.add_argument("input")
    dedupe_p.add_argument("--config", required=True)
    dedupe_p.set_defaults(func=cmd_dedupe)

    restore_p = sub.add_parser("restore")
    restore_p.add_argument("input")
    restore_p.add_argument("--config", required=True)
    restore_p.set_defaults(func=cmd_restore)

    sub.add_parser("checksum").set_defaults(func=cmd_checksum)
    sub.add_parser("write-approved-checksums").set_defaults(func=cmd_write_approved_checksums)
    sub.add_parser("verify-self").set_defaults(func=cmd_verify_self)

    validate_p = sub.add_parser("validate-config")
    validate_p.add_argument("config")
    validate_p.set_defaults(func=cmd_validate_config)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        args.func(args)
        return 0
    except PrivacyToolError as exc:
        safe_exit(str(exc))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

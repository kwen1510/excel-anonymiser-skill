import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
TOOL = ROOT / "privacy_handler" / "excel_privacy_tool.py"
SENSITIVE = [
    "Alice Example",
    "Bob Example",
    "Class A",
    "Class B",
    "alice@example.test",
    "bob@example.test",
    "mock-password",
]


@pytest.fixture()
def isolated_project(tmp_path):
    for name in ["privacy_handler", "configs"]:
        src = ROOT / name
        dst = tmp_path / name
        dst.mkdir()
        for item in src.iterdir():
            if item.is_file():
                dst.joinpath(item.name).write_bytes(item.read_bytes())
    for name in ["original_data", "safe_for_codex", "private_keys", "restored_outputs"]:
        tmp_path.joinpath(name).mkdir()
        tmp_path.joinpath(name, ".gitkeep").write_text("")
    return tmp_path


def run_tool(project, *args, check=True):
    result = subprocess.run(
        [sys.executable, str(project / "privacy_handler" / "excel_privacy_tool.py"), *args],
        cwd=project,
        text=True,
        capture_output=True,
        check=False,
    )
    if check and result.returncode != 0:
        raise AssertionError(f"Command failed: {result.stderr}")
    return result


def create_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Responses"
    ws.append(["Name", "Email", "Class", "Marks", "Comments", "QuestionID"])
    ws.append(["Alice Example", "alice@example.test", "Class A", 10, "Keep this", "Q1"])
    ws.append(["Bob Example", "bob@example.test", "Class B", 8, "Keep that", "Q1"])
    ws.append(["Alice Example", "alice@example.test", "Class A", 10, None, "Q1"])
    wb.save(path)


def privacy_config(project, **overrides):
    config = {
        "header_row": 1,
        "anonymise_sheet_names": False,
        "anonymise_headers": False,
        "output": {
            "safe_dir": "safe_for_codex",
            "private_key_dir": "private_keys",
            "restored_dir": "restored_outputs",
        },
        "key_file": {
            "format": "xlsx",
            "encrypt_original_values": False,
            "secret_key_file": None,
        },
        "sheets": {
            "Responses": {
                "columns": {
                    "Name": {"mode": "uuid", "prefix": "NAME"},
                    "Email": {"mode": "uuid", "prefix": "EMAIL"},
                    "Class": {"mode": "uuid", "prefix": "CLASS"},
                    "Marks": {"mode": "keep"},
                    "Comments": {"mode": "keep"},
                }
            }
        },
    }
    config.update(overrides)
    path = project / "configs" / "privacy_config.json"
    path.write_text(json.dumps(config), encoding="utf-8")
    return path


def approve(project):
    run_tool(project, "write-approved-checksums")
    result = run_tool(project, "verify-self")
    assert "passed" in result.stdout


def assert_no_sensitive_output(result):
    combined = result.stdout + result.stderr
    for value in SENSITIVE:
        assert value not in combined


def test_inspect_returns_safe_metadata_only(isolated_project):
    project = isolated_project
    source = project / "original_data" / "original.xlsx"
    create_workbook(source)

    result = run_tool(project, "inspect", "original_data/original.xlsx", "--header-row", "1")
    data = json.loads(result.stdout)

    assert data["workbook"] == "original.xlsx"
    assert data["sheets"][0]["sheet_name"] == "Responses"
    assert data["sheets"][0]["rows"] == 4
    assert data["sheets"][0]["columns"] == 6
    assert [h["header"] for h in data["sheets"][0]["headers"]] == [
        "Name",
        "Email",
        "Class",
        "Marks",
        "Comments",
        "QuestionID",
    ]
    assert_no_sensitive_output(result)


def test_anonymise_creates_safe_outputs_and_uuid_mappings(isolated_project):
    project = isolated_project
    source = project / "original_data" / "original.xlsx"
    create_workbook(source)
    privacy_config(project)
    approve(project)

    result = run_tool(project, "anonymise", "original_data/original.xlsx", "--config", "configs/privacy_config.json")
    assert_no_sensitive_output(result)

    safe = project / "safe_for_codex" / "original.anonymised.xlsx"
    manifest = project / "safe_for_codex" / "original.privacy_manifest.json"
    key = project / "private_keys" / "original.privacy_key.xlsx"
    assert safe.exists()
    assert manifest.exists()
    assert key.exists()
    assert not (project / "safe_for_codex" / "original.privacy_key.xlsx").exists()

    wb = load_workbook(safe)
    ws = wb["Responses"]
    alice_1 = ws["A2"].value
    bob = ws["A3"].value
    alice_2 = ws["A4"].value
    assert alice_1 == alice_2
    assert alice_1 != bob
    assert alice_1.startswith("NAME_")
    assert len(alice_1) == len("NAME_550e8400")
    assert ws["E4"].value is None

    manifest_text = manifest.read_text(encoding="utf-8")
    for value in SENSITIVE:
        assert value not in manifest_text


def test_restore_recreates_synthetic_original_values(isolated_project):
    project = isolated_project
    source = project / "original_data" / "original.xlsx"
    create_workbook(source)
    privacy_config(project)
    approve(project)
    run_tool(project, "anonymise", "original_data/original.xlsx", "--config", "configs/privacy_config.json")

    restore_config = {
        "key_file": "private_keys/original.privacy_key.xlsx",
        "secret_key_file": None,
        "output_file": "restored_outputs/original.restored.xlsx",
    }
    (project / "configs" / "restore_config.json").write_text(json.dumps(restore_config), encoding="utf-8")
    result = run_tool(
        project,
        "restore",
        "safe_for_codex/original.anonymised.xlsx",
        "--config",
        "configs/restore_config.json",
    )
    assert_no_sensitive_output(result)
    wb = load_workbook(project / "restored_outputs" / "original.restored.xlsx")
    ws = wb["Responses"]
    assert ws["A2"].value == "Alice Example"
    assert ws["A3"].value == "Bob Example"
    assert ws["B2"].value == "alice@example.test"
    assert ws["C3"].value == "Class B"


def test_simple_mask_and_unmask_with_encrypted_csv_key(isolated_project):
    project = isolated_project
    source = project / "original_data" / "original.xlsx"
    create_workbook(source)
    approve(project)

    result = run_tool(
        project,
        "mask",
        "original_data/original.xlsx",
        "--columns",
        "Name,Email,Class",
        "--sheet",
        "Responses",
    )
    assert_no_sensitive_output(result)
    masked = project / "safe_for_codex" / "original.masked.xlsx"
    key = project / "private_keys" / "original.mask_key.csv"
    secret = project / "private_keys" / "original.mask_secret.key"
    assert masked.exists()
    assert key.exists()
    assert secret.exists()

    masked_wb = load_workbook(masked)
    masked_ws = masked_wb["Responses"]
    assert masked_ws["A2"].value.startswith("MASK_")
    assert masked_ws["A2"].value == masked_ws["A4"].value
    assert masked_ws["A2"].value != masked_ws["A3"].value

    key_text = key.read_text(encoding="utf-8")
    for value in SENSITIVE:
        assert value not in key_text

    result = run_tool(
        project,
        "unmask",
        "safe_for_codex/original.masked.xlsx",
        "--key-file",
        "private_keys/original.mask_key.csv",
        "--secret-key-file",
        "private_keys/original.mask_secret.key",
        "--output",
        "restored_outputs/original.unmasked.xlsx",
    )
    assert_no_sensitive_output(result)
    restored = load_workbook(project / "restored_outputs" / "original.unmasked.xlsx")
    ws = restored["Responses"]
    assert ws["A2"].value == "Alice Example"
    assert ws["A3"].value == "Bob Example"
    assert ws["B2"].value == "alice@example.test"
    assert ws["C3"].value == "Class B"


def test_dedupe_removes_duplicate_rows(isolated_project):
    project = isolated_project
    source = project / "safe_for_codex" / "processed.xlsx"
    create_workbook(source)
    approve(project)
    config = {
        "sheet": "Responses",
        "by": ["Name", "Class", "QuestionID"],
        "keep": "first",
        "output_file": "safe_for_codex/processed.deduped.xlsx",
    }
    (project / "configs" / "dedupe_config.json").write_text(json.dumps(config), encoding="utf-8")
    result = run_tool(project, "dedupe", "safe_for_codex/processed.xlsx", "--config", "configs/dedupe_config.json")
    assert_no_sensitive_output(result)
    assert "Rows before: 3" in result.stdout
    assert "Rows after: 2" in result.stdout


def test_missing_configured_column_error_is_safe(isolated_project):
    project = isolated_project
    source = project / "original_data" / "original.xlsx"
    create_workbook(source)
    config_path = privacy_config(project)
    config = json.loads(config_path.read_text())
    config["sheets"]["Responses"]["columns"]["Missing"] = {"mode": "uuid", "prefix": "MISS"}
    config_path.write_text(json.dumps(config), encoding="utf-8")
    approve(project)

    result = run_tool(
        project,
        "anonymise",
        "original_data/original.xlsx",
        "--config",
        "configs/privacy_config.json",
        check=False,
    )
    assert result.returncode != 0
    assert "Missing configured column" in result.stderr
    assert_no_sensitive_output(result)


def test_header_and_sheet_name_anonymisation(isolated_project):
    project = isolated_project
    source = project / "original_data" / "original.xlsx"
    create_workbook(source)
    privacy_config(project, anonymise_sheet_names=True, anonymise_headers=True)
    approve(project)
    run_tool(project, "anonymise", "original_data/original.xlsx", "--config", "configs/privacy_config.json")

    wb = load_workbook(project / "safe_for_codex" / "original.anonymised.xlsx")
    assert wb.sheetnames[0].startswith("SHEET_")
    ws = wb[wb.sheetnames[0]]
    assert ws["A1"].value.startswith("HEADER_")


def test_restore_after_header_and_sheet_name_anonymisation(isolated_project):
    project = isolated_project
    source = project / "original_data" / "original.xlsx"
    create_workbook(source)
    privacy_config(project, anonymise_sheet_names=True, anonymise_headers=True)
    approve(project)
    run_tool(project, "anonymise", "original_data/original.xlsx", "--config", "configs/privacy_config.json")

    restore_config = {
        "key_file": "private_keys/original.privacy_key.xlsx",
        "secret_key_file": None,
        "output_file": "restored_outputs/original.restored.xlsx",
    }
    (project / "configs" / "restore_config.json").write_text(json.dumps(restore_config), encoding="utf-8")
    result = run_tool(
        project,
        "restore",
        "safe_for_codex/original.anonymised.xlsx",
        "--config",
        "configs/restore_config.json",
    )
    assert_no_sensitive_output(result)
    wb = load_workbook(project / "restored_outputs" / "original.restored.xlsx")
    ws = wb[wb.sheetnames[0]]
    assert ws["A2"].value == "Alice Example"
    assert ws["B3"].value == "bob@example.test"


def test_checksum_and_validate_config(isolated_project):
    project = isolated_project
    privacy_config(project)
    result = run_tool(project, "checksum")
    checksums = json.loads(result.stdout)
    assert "privacy_handler/excel_privacy_tool.py" in checksums

    result = run_tool(project, "validate-config", "configs/privacy_config.json")
    assert "Config valid" in result.stdout

    approve(project)


def test_validate_config_rejects_unsafe_private_output_paths(isolated_project):
    project = isolated_project
    config_path = privacy_config(project)
    config = json.loads(config_path.read_text())
    config["output"]["private_key_dir"] = "safe_for_codex/private"
    config_path.write_text(json.dumps(config), encoding="utf-8")

    result = run_tool(project, "validate-config", "configs/privacy_config.json", check=False)
    assert result.returncode != 0
    assert "Private key directory" in result.stderr
    assert_no_sensitive_output(result)


def test_password_provider_path_does_not_expose_password(monkeypatch, isolated_project):
    project = isolated_project
    sys.path.insert(0, str(project / "privacy_handler"))
    import excel_privacy_tool as tool

    source = project / "original_data" / "original.xlsx"
    source.write_bytes(b"not an xlsx")

    class FakeOfficeFile:
        def __init__(self, _file):
            pass

        def load_key(self, password):
            assert password == "mock-password"

        def decrypt(self, _out):
            raise RuntimeError("synthetic decrypt failure")

    class FakeMsoffcrypto:
        OfficeFile = FakeOfficeFile

    monkeypatch.setattr(tool, "msoffcrypto", FakeMsoffcrypto)
    with pytest.raises(tool.PrivacyToolError) as exc:
        tool.load_workbook_private(source, password_provider=lambda _name: "mock-password")
    assert "mock-password" not in str(exc.value)
